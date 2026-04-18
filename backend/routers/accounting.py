"""
Роутер для управления бухгалтерией мастера
"""
import logging
from datetime import datetime, timedelta, date
from typing import Optional, List
from io import BytesIO
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from fastapi import APIRouter, Depends, HTTPException, Query, Response, status
from sqlalchemy.orm import Session
from sqlalchemy import func, desc, and_, or_

from database import get_db
from models import (
    User, Master, IndieMaster, MasterExpense, BookingConfirmation, Booking, Income,
    Service, BookingStatus, TaxRate, MasterClientMetadata
)
from auth import get_current_active_user
from utils.client_display_name import get_client_display_name, get_meta_for_client, strip_indie_service_prefix
from utils.subscription_features import has_finance_access

router = APIRouter(prefix="/api/master/accounting", tags=["accounting"])


def _ensure_finance_access(db: Session, user_id: int) -> None:
    """Проверка доступа к разделу «Финансы»; 403 при отсутствии."""
    if not has_finance_access(db, user_id):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Finance feature not available in your plan",
            headers={"X-Error-Code": "FEATURE_NOT_AVAILABLE"},
        )


def _booking_status_str(booking: Booking) -> str:
    s = booking.status
    if s is None:
        return ""
    return s.value if hasattr(s, "value") else str(s)


def _is_future_pre_visit_to_confirmed(new_status: str, booking: Booking) -> bool:
    """
    Транзишен подтверждения будущей записи (booking workflow, не «Финансы»):
    future + (created | awaiting_confirmation) → confirmed.
    """
    if new_status != "confirmed":
        return False
    now_u = datetime.utcnow()
    if booking.start_time <= now_u:
        return False
    return _booking_status_str(booking).lower() in ("created", "awaiting_confirmation")


logger = logging.getLogger(__name__)


@router.post("/update-booking-status/{booking_id}")
async def update_booking_status(
    booking_id: int,
    new_status: str = Query(..., regex="^(created|confirmed|awaiting_confirmation|completed|cancelled|client_requested_early|client_requested_late)$"),
    cancellation_reason: str = Query(None, regex="^(client_requested|client_no_show|mutual_agreement|master_unavailable)$"),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Изменить статус прошедшей записи (редактирование ошибки мастера).
    Разрешенные значения: created, awaiting_confirmation, completed, cancelled.
    При изменении статуса:
    - Если переводим в completed: создаем BookingConfirmation и Income (если еще нет)
    - Если переводим из completed в другой: удаляем BookingConfirmation и связанные доходы
    - Для cancelled: просто меняем статус
    """
    try:
        # IMPORTANT:
        # - Booking.master_id / indie_master_id — проверка владения (салон / инди)
        # - BookingConfirmation/MasterExpense/TaxRate.master_id хранит users.id (финансы мастера)
        master_user_id = current_user.id
        master_id, indie_id = get_booking_owner_ids(current_user.id, db)
        owner_cond = _booking_owner_filter(Booking, master_id, indie_id)

        booking = db.query(Booking).filter(
            Booking.id == booking_id,
            owner_cond
        ).first()
        if not booking:
            raise HTTPException(status_code=404, detail="Бронирование не найдено")

        # Ручное подтверждение будущей записи — часть booking workflow, не «Финансы».
        if not _is_future_pre_visit_to_confirmed(new_status, booking):
            _ensure_finance_access(db, current_user.id)

        # Pre-visit: future + (created | awaiting_confirmation) → confirmed — только ручной режим мастера.
        if new_status == "confirmed":
            now_u = datetime.utcnow()
            st = _booking_status_str(booking).lower()
            if booking.start_time > now_u and st in ("created", "awaiting_confirmation"):
                master_row = db.query(Master).filter(Master.user_id == current_user.id).first()
                if not master_row:
                    raise HTTPException(
                        status_code=status.HTTP_403_FORBIDDEN,
                        detail="Профиль мастера не найден.",
                    )
                manual = (
                    master_row.auto_confirm_bookings is False
                    or master_row.auto_confirm_bookings is None
                )
                if not manual:
                    raise HTTPException(
                        status_code=status.HTTP_403_FORBIDDEN,
                        detail="При автоматическом подтверждении новых записей ручное принятие не требуется.",
                    )

        # awaiting_confirmation — только для post-visit (прошлые записи).
        # Для будущих записей запрещено выставлять awaiting_confirmation.
        if new_status == "awaiting_confirmation" and booking.start_time > datetime.utcnow():
            raise HTTPException(
                status_code=400,
                detail="Статус 'На подтверждение' применим только к прошедшим записям. Для будущих используйте 'confirmed'.",
            )

        # Проверяем, что запись не отменена клиентом (мастер не может изменять такие записи)
        if booking.status in [BookingStatus.CANCELLED_BY_CLIENT_EARLY, BookingStatus.CANCELLED_BY_CLIENT_LATE]:
            raise HTTPException(
                status_code=403, 
                detail="Нельзя изменять статус записи, отмененной клиентом"
            )

        # Очистка подтверждений/доходов, если уходим из completed
        if booking.status == BookingStatus.COMPLETED and new_status != "completed":
            confirmation = db.query(BookingConfirmation).filter(BookingConfirmation.booking_id == booking_id).first()
            if confirmation:
                db.delete(confirmation)
            income = db.query(Income).filter(Income.booking_id == booking_id).first()
            if income:
                db.delete(income)

        # Установка нового статуса
        if new_status == "created":
            booking.status = BookingStatus.CREATED
        elif new_status == "confirmed":
            booking.status = BookingStatus.CONFIRMED
        elif new_status == "awaiting_confirmation":
            booking.status = BookingStatus.AWAITING_CONFIRMATION
        elif new_status == "cancelled":
            booking.status = BookingStatus.CANCELLED
            if cancellation_reason:
                booking.cancellation_reason = cancellation_reason
                booking.cancelled_by_user_id = current_user.id
        elif new_status == "client_requested_early":
            booking.status = BookingStatus.CANCELLED_BY_CLIENT_EARLY
            booking.cancellation_reason = "client_requested"
            booking.cancelled_by_user_id = booking.client_id  # Клиент отменил сам
        elif new_status == "client_requested_late":
            booking.status = BookingStatus.CANCELLED_BY_CLIENT_LATE
            booking.cancellation_reason = "client_requested"
            booking.cancelled_by_user_id = booking.client_id  # Клиент отменил сам
        elif new_status == "completed":
            booking.status = BookingStatus.COMPLETED
            # Создаем подтверждение и доход при необходимости
            existing_confirmation = db.query(BookingConfirmation).filter(
                BookingConfirmation.booking_id == booking_id
            ).first()
            if not existing_confirmation:
                confirmation = BookingConfirmation(
                    booking_id=booking_id,
                    master_id=master_user_id,
                    confirmed_income=booking.payment_amount or 0
                )
                db.add(confirmation)

            existing_income = db.query(Income).filter(Income.booking_id == booking_id).first()
            if not existing_income:
                income = Income(
                    # Income.indie_master_id -> FK на indie_masters.id (не masters.id)
                    indie_master_id=None,
                    booking_id=booking_id,
                    total_amount=booking.payment_amount or 0,
                    master_earnings=booking.payment_amount or 0,
                    salon_earnings=0,
                    income_date=datetime.utcnow().date(),
                    service_date=booking.start_time.date()
                )
                db.add(income)

        db.commit()
        return {"message": "Статус записи обновлен", "booking_id": booking_id, "new_status": new_status}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка при изменении статуса записи {booking_id}: {e}")
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

def get_master_id_from_user(user_id: int, db: Session) -> int:
    """Получить Master.id (masters.id) по User.id (users.id)."""
    master = db.query(Master).filter(Master.user_id == user_id).first()
    if not master:
        raise HTTPException(status_code=404, detail="Профиль мастера не найден")
    return master.id


def get_booking_owner_ids(user_id: int, db: Session) -> tuple[Optional[int], Optional[int]]:
    """Возвращает (master_id, indie_master_id) для проверки владения Booking.
    Booking имеет либо master_id (салон), либо indie_master_id (инди)."""
    master = db.query(Master).filter(Master.user_id == user_id).first()
    indie = db.query(IndieMaster).filter(IndieMaster.user_id == user_id).first()
    master_id = master.id if master else None
    indie_id = indie.id if indie else None
    if not master_id and not indie_id:
        raise HTTPException(status_code=404, detail="Профиль мастера не найден")
    return (master_id, indie_id)


def _booking_owner_filter(Booking, master_id: Optional[int], indie_id: Optional[int]):
    """Условие владения: master_id или indie_master_id."""
    conds = []
    if master_id is not None:
        conds.append(Booking.master_id == master_id)
    if indie_id is not None:
        conds.append(Booking.indie_master_id == indie_id)
    if not conds:
        return and_(False)  # no match
    return or_(*conds)


def auto_confirm_awaiting_on_manual_switch(db: Session, user_id: int) -> int:
    """При переключении на ручной режим: автоподтвердить все текущие AWAITING_CONFIRMATION.
    Возвращает количество подтверждённых записей."""
    master_row_id = get_master_id_from_user(user_id, db)
    master_user_id = user_id
    master_id, indie_id = get_booking_owner_ids(user_id, db)
    owner_cond = _booking_owner_filter(Booking, master_id, indie_id)

    pending_bookings = db.query(Booking).outerjoin(
        BookingConfirmation,
        Booking.id == BookingConfirmation.booking_id
    ).filter(
        owner_cond,
        Booking.status == BookingStatus.AWAITING_CONFIRMATION,
        BookingConfirmation.id == None
    ).all()

    confirmed_count = 0
    for booking in pending_bookings:
        confirmation = BookingConfirmation(
            booking_id=booking.id,
            master_id=master_user_id,
            confirmed_income=booking.payment_amount or 0
        )
        db.add(confirmation)
        booking.status = BookingStatus.COMPLETED
        confirmed_count += 1

    if confirmed_count:
        logger.info(f"[manual-switch] Автоподтверждено {confirmed_count} записей для мастера user_id={user_id}")
    return confirmed_count


def get_tax_rate_for_date(master_id: int, target_date: datetime, db: Session) -> float:
    """Получить налоговую ставку для конкретной даты.

    NOTE: TaxRate.master_id хранит users.id.
    """
    tax_rate = db.query(TaxRate).filter(
        TaxRate.master_id == master_id,
        TaxRate.effective_from_date <= target_date.date()
    ).order_by(desc(TaxRate.effective_from_date)).first()
    
    return tax_rate.rate if tax_rate else 0.0


def resolve_accounting_calendar_bounds(
    period: str,
    offset: int = 0,
    anchor_date: Optional[str] = None,
    window_before: Optional[int] = None,
    window_after: Optional[int] = None,
) -> tuple[datetime, datetime]:
    """
    Границы календарного периода для «Финансы» — те же, что у GET /api/master/dashboard/stats
    (utils.master_stats_periods.build_stats_periods_bundle, bucket is_current).
    """
    from utils.master_stats_periods import build_stats_periods_bundle

    _, _, cur_pd, _ = build_stats_periods_bundle(
        period,
        offset,
        anchor_date,
        window_before,
        window_after,
    )
    if not cur_pd:
        raise HTTPException(status_code=400, detail="Не удалось вычислить границы периода")
    start_date = datetime.combine(cur_pd["start"], datetime.min.time())
    end_date = datetime.combine(cur_pd["end"], datetime.max.time())
    return start_date, end_date


@router.get("/expenses")
async def get_expenses(
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    expense_type: Optional[str] = None,
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Получить список расходов мастера с фильтрами и пагинацией"""
    _ensure_finance_access(db, current_user.id)
    try:
        master_user_id = current_user.id
        query = db.query(MasterExpense).filter(MasterExpense.master_id == master_user_id)
        
        if start_date:
            query = query.filter(MasterExpense.expense_date >= start_date)
        if end_date:
            query = query.filter(MasterExpense.expense_date <= end_date)
        if expense_type:
            query = query.filter(MasterExpense.expense_type == expense_type)
        
        total = query.count()
        expenses = query.order_by(desc(MasterExpense.expense_date)).offset((page - 1) * limit).limit(limit).all()
        
        return {
            "expenses": [
                {
                    "id": exp.id,
                    "name": exp.name,
                    "expense_type": exp.expense_type,
                    "amount": exp.amount,
                    "recurrence_type": exp.recurrence_type,
                    "condition_type": exp.condition_type,
                    "service_id": exp.service_id,
                    "expense_date": exp.expense_date,
                    "is_active": exp.is_active,
                    "created_at": exp.created_at,
                }
                for exp in expenses
            ],
            "total": total,
            "page": page,
            "limit": limit,
            "pages": (total + limit - 1) // limit
        }
    except Exception as e:
        logger.error(f"Ошибка при получении расходов: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/operations")
async def get_operations(
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    period: Optional[str] = Query(
        None,
        regex="^(day|week|month|quarter|year)$",
        description="Если задан без start/end — те же границы, что у /accounting/summary",
    ),
    offset: int = Query(0),
    anchor_date: Optional[str] = Query(None),
    window_before: Optional[int] = Query(None, ge=0, le=31),
    window_after: Optional[int] = Query(None, ge=0, le=31),
    operation_type: Optional[str] = None,  # "income", "expense", "all"
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Получить все операции мастера (доходы + расходы) с фильтрами и пагинацией"""
    _ensure_finance_access(db, current_user.id)
    try:
        operations = []
        master_user_id = current_user.id

        if (start_date is None or end_date is None) and period is not None:
            start_date, end_date = resolve_accounting_calendar_bounds(
                period,
                offset,
                anchor_date,
                window_before,
                window_after,
            )

        # Получаем доходы (подтвержденные)
        income_query = db.query(BookingConfirmation).filter(
            BookingConfirmation.master_id == master_user_id
        )
        
        if start_date:
            income_query = income_query.filter(BookingConfirmation.confirmed_at >= start_date)
        if end_date:
            income_query = income_query.filter(BookingConfirmation.confirmed_at <= end_date)
        
        if operation_type != "expense":
            incomes = income_query.order_by(desc(BookingConfirmation.confirmed_at)).all()
            for income in incomes:
                # Получаем налоговую ставку для даты подтверждения
                tax_rate = get_tax_rate_for_date(master_user_id, income.confirmed_at, db)
                gross_amount = income.confirmed_income
                tax_amount = gross_amount * (tax_rate / 100)
                net_amount = gross_amount - tax_amount
                
                operations.append({
                    "id": f"income_{income.id}",
                    "date": income.confirmed_at,
                    "name": f"Доход от услуги",
                    "type": "Доход",
                    "operation_type": "income",
                    "amount": net_amount,  # Чистый доход с учетом налога
                    "gross_amount": gross_amount,
                    "tax_rate": tax_rate,
                    "tax_amount": tax_amount,
                    "net_amount": net_amount,
                    "description": f"Подтвержденный доход от бронирования #{income.booking_id}"
                })
        
        # Получаем расходы
        expense_query = db.query(MasterExpense).filter(
            MasterExpense.master_id == master_user_id,
            MasterExpense.is_active == True
        )
        
        if start_date:
            expense_query = expense_query.filter(MasterExpense.expense_date >= start_date)
        if end_date:
            expense_query = expense_query.filter(MasterExpense.expense_date <= end_date)
        
        if operation_type != "income":
            expenses = expense_query.order_by(desc(MasterExpense.expense_date)).all()
            for expense in expenses:
                operations.append({
                    "id": f"expense_{expense.id}",
                    "date": expense.expense_date,
                    "name": expense.name,
                    "type": "Расход",
                    "operation_type": "expense",
                    "amount": -expense.amount,  # Отрицательное значение для расходов
                    "description": f"{expense.expense_type} - {expense.name}"
                })
        
        # Сортируем по дате (новые сначала), обрабатываем None
        operations.sort(key=lambda x: x["date"] if x["date"] is not None else datetime.min, reverse=True)
        
        # Применяем пагинацию
        total = len(operations)
        start_idx = (page - 1) * limit
        end_idx = start_idx + limit
        paginated_operations = operations[start_idx:end_idx]
        
        return {
            "operations": paginated_operations,
            "total": total,
            "page": page,
            "limit": limit,
            "pages": (total + limit - 1) // limit
        }
    except Exception as e:
        logger.error(f"Ошибка при получении операций: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/expenses")
async def create_expense(
    name: str,
    expense_type: str,
    amount: float,
    recurrence_type: Optional[str] = None,
    condition_type: Optional[str] = None,
    service_id: Optional[int] = None,
    expense_date: Optional[datetime] = None,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Создать новый расход"""
    _ensure_finance_access(db, current_user.id)
    try:
        # Валидация в зависимости от типа расхода
        if expense_type == "recurring" and not recurrence_type:
            raise HTTPException(status_code=400, detail="Для циклического расхода требуется recurrence_type")
        if expense_type == "service_based" and not service_id:
            raise HTTPException(status_code=400, detail="Для расхода по услуге требуется service_id")
        if expense_type == "one_time" and not expense_date:
            raise HTTPException(status_code=400, detail="Для разового расхода требуется expense_date")
        
        master_user_id = current_user.id
        
        expense = MasterExpense(
            master_id=master_user_id,
            name=name,
            expense_type=expense_type,
            amount=amount,
            recurrence_type=recurrence_type,
            condition_type=condition_type,
            service_id=service_id,
            expense_date=expense_date or datetime.utcnow()
        )
        
        db.add(expense)
        db.commit()
        db.refresh(expense)
        
        logger.info(f"Создан расход {expense.id} для мастера {current_user.id}")
        
        return {
            "id": expense.id,
            "name": expense.name,
            "expense_type": expense.expense_type,
            "amount": expense.amount,
            "message": "Расход успешно создан"
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка при создании расхода: {e}")
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/expenses/{expense_id}")
async def update_expense(
    expense_id: int,
    name: Optional[str] = None,
    expense_type: Optional[str] = None,
    amount: Optional[float] = None,
    recurrence_type: Optional[str] = None,
    condition_type: Optional[str] = None,
    service_id: Optional[int] = None,
    expense_date: Optional[datetime] = None,
    is_active: Optional[bool] = None,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Редактировать расход"""
    _ensure_finance_access(db, current_user.id)
    try:
        master_user_id = current_user.id
        expense = db.query(MasterExpense).filter(
            MasterExpense.id == expense_id,
            MasterExpense.master_id == master_user_id
        ).first()
        
        if not expense:
            raise HTTPException(status_code=404, detail="Расход не найден")
        
        if name is not None:
            expense.name = name
        if expense_type is not None:
            expense.expense_type = expense_type
        if amount is not None:
            expense.amount = amount
        if recurrence_type is not None:
            expense.recurrence_type = recurrence_type
        if condition_type is not None:
            expense.condition_type = condition_type
        if service_id is not None:
            expense.service_id = service_id
        if expense_date is not None:
            expense.expense_date = expense_date
        if is_active is not None:
            expense.is_active = is_active
        
        expense.updated_at = datetime.utcnow()
        db.commit()
        
        logger.info(f"Обновлен расход {expense_id} для мастера {current_user.id}")
        
        return {"message": "Расход успешно обновлен"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка при обновлении расхода: {e}")
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/expenses/{expense_id}")
async def delete_expense(
    expense_id: int,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Удалить расход"""
    _ensure_finance_access(db, current_user.id)
    try:
        master_user_id = current_user.id
        expense = db.query(MasterExpense).filter(
            MasterExpense.id == expense_id,
            MasterExpense.master_id == master_user_id
        ).first()
        
        if not expense:
            raise HTTPException(status_code=404, detail="Расход не найден")
        
        db.delete(expense)
        db.commit()
        
        logger.info(f"Удален расход {expense_id} для мастера {current_user.id}")
        
        return {"message": "Расход успешно удален"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка при удалении расхода: {e}")
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


def _accounting_chart_granularity(period: str, day_count: int, explicit_date_range: bool) -> str:
    """Гранулярность точек графика: согласована с выбранным пресетом / длиной custom range."""
    if explicit_date_range:
        if day_count <= 31:
            return "day"
        if day_count <= 93:
            return "week"
        return "month"
    if period == "day":
        return "day"
    if period == "week":
        # Динамика внутри выбранной недели — по дням (ровно окно start/end сводки)
        return "day"
    if period == "month":
        return "week"
    if period in ("quarter", "year"):
        return "month"
    return "day"


def _accounting_fill_daily_chart_rows(
    start_dt: datetime,
    end_dt: datetime,
    income_by_date: dict,
    expected_income_by_date: dict,
    expenses_by_date: dict,
) -> list[dict]:
    rows: list[dict] = []
    cur = start_dt.date()
    end_d = end_dt.date()
    while cur <= end_d:
        ds = cur.isoformat()
        inc = float(income_by_date.get(ds, 0) or 0)
        exp_inc = float(expected_income_by_date.get(ds, 0) or 0)
        ex = float(expenses_by_date.get(ds, 0) or 0)
        rows.append(
            {
                "date": ds,
                "income": inc,
                "expected_income": exp_inc,
                "expense": ex,
                "net_profit": inc - ex,
            }
        )
        cur += timedelta(days=1)
    return rows


def _accounting_rollup_chart_rows(rows: list[dict], granularity: str) -> list[dict]:
    if granularity == "day" or len(rows) <= 1:
        return rows
    buckets: dict[tuple, dict] = {}
    for row in rows:
        d = datetime.strptime(row["date"], "%Y-%m-%d").date()
        if granularity == "week":
            iso = d.isocalendar()
            key = (iso.year, iso.week)
            label = date.fromisocalendar(iso.year, iso.week, 1).isoformat()
        elif granularity == "month":
            key = (d.year, d.month)
            label = date(d.year, d.month, 1).isoformat()
        else:
            return rows
        if key not in buckets:
            buckets[key] = {
                "date": label,
                "income": 0.0,
                "expected_income": 0.0,
                "expense": 0.0,
                "net_profit": 0.0,
            }
        buckets[key]["income"] += float(row["income"] or 0)
        buckets[key]["expected_income"] += float(row["expected_income"] or 0)
        buckets[key]["expense"] += float(row["expense"] or 0)
        buckets[key]["net_profit"] += float(row["net_profit"] or 0)
    return sorted(buckets.values(), key=lambda x: x["date"])


@router.get("/summary")
async def get_accounting_summary(
    period: str = Query("week", regex="^(day|week|month|quarter|year)$"),
    offset: int = Query(0),
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    anchor_date: Optional[str] = Query(
        None,
        description="YYYY-MM-DD для period=day — как у /api/master/dashboard/stats",
    ),
    window_before: Optional[int] = Query(None, ge=0, le=31),
    window_after: Optional[int] = Query(None, ge=0, le=31),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Получить сводку доходов/расходов с данными для графиков"""
    _ensure_finance_access(db, current_user.id)
    try:
        had_explicit_date_range = start_date is not None and end_date is not None
        # Определяем даты периода (единый контракт с dashboard/stats)
        if start_date and end_date:
            # Свободный выбор диапазона
            pass
        else:
            start_date, end_date = resolve_accounting_calendar_bounds(
                period,
                offset,
                anchor_date,
                window_before,
                window_after,
            )

        logger.info(f"Получение сводки за период {period} (offset={offset}), даты: {start_date} - {end_date}")
        
        master_row_id = get_master_id_from_user(current_user.id, db)
        master_user_id = current_user.id
        booking_master_id, booking_indie_id = get_booking_owner_ids(current_user.id, db)
        owner_expected = _booking_owner_filter(Booking, booking_master_id, booking_indie_id)

        # Получаем подтвержденные доходы (детально для расчета налога)
        confirmed_incomes_detailed = db.query(BookingConfirmation).filter(
            BookingConfirmation.master_id == master_user_id,
            BookingConfirmation.confirmed_at >= start_date,
            BookingConfirmation.confirmed_at <= end_date
        ).all()
        
        # Группируем по датам и рассчитываем налог
        income_by_date = {}
        points_spent_by_date = {}  # Для учета списанных баллов
        total_points_spent = 0
        
        for income in confirmed_incomes_detailed:
            date_str = str(income.confirmed_at.date())
            tax_rate = get_tax_rate_for_date(master_user_id, income.confirmed_at, db)
            net_amount = income.confirmed_income * (1 - tax_rate / 100)
            
            if date_str not in income_by_date:
                income_by_date[date_str] = 0
            income_by_date[date_str] += net_amount
            
            # Получаем информацию о списанных баллах из booking
            if income.booking_id:
                booking = db.query(Booking).filter(Booking.id == income.booking_id).first()
                if booking and booking.loyalty_points_used and booking.loyalty_points_used > 0:
                    points_spent = booking.loyalty_points_used
                    if date_str not in points_spent_by_date:
                        points_spent_by_date[date_str] = 0
                    points_spent_by_date[date_str] += points_spent
                    total_points_spent += points_spent
        
        # Ожидаемые доходы: ещё без строки BookingConfirmation (как pending-стек dashboard/stats:
        # created / confirmed / awaiting_confirmation).
        expected_incomes = db.query(
            func.sum(Booking.payment_amount).label("total_expected_income"),
            func.date(Booking.start_time).label("date")
        ).filter(
            owner_expected,
            Booking.start_time >= start_date,
            Booking.start_time <= end_date,
            Booking.status.in_(
                [
                    BookingStatus.CREATED,
                    BookingStatus.CONFIRMED,
                    BookingStatus.AWAITING_CONFIRMATION,
                ]
            ),
        ).group_by(func.date(Booking.start_time)).all()
        
        # Получаем расходы
        expenses_query = db.query(
            func.sum(MasterExpense.amount).label("total_expense"),
            func.date(MasterExpense.expense_date).label("date")
        ).filter(
            MasterExpense.master_id == master_user_id,
            MasterExpense.expense_date >= start_date,
            MasterExpense.expense_date <= end_date,
            MasterExpense.is_active == True
        ).group_by(func.date(MasterExpense.expense_date)).all()
        
        # Формируем данные для графиков
        expected_income_by_date = {str(row.date): float(row.total_expected_income or 0) for row in expected_incomes}
        expenses_by_date = {str(row.date): float(row.total_expense or 0) for row in expenses_query}
        
        # Точки графика: полный календарный ряд по окну сводки + rollup по гранулярности пресета/custom
        all_dates = sorted(
            set(
                list(income_by_date.keys())
                + list(expected_income_by_date.keys())
                + list(expenses_by_date.keys())
            )
        )
        chart_axis_granularity = "day"
        chart_data: list[dict] = []
        if all_dates:
            day_count = (end_date.date() - start_date.date()).days + 1
            chart_axis_granularity = _accounting_chart_granularity(
                period, day_count, had_explicit_date_range
            )
            daily_rows = _accounting_fill_daily_chart_rows(
                start_date,
                end_date,
                income_by_date,
                expected_income_by_date,
                expenses_by_date,
            )
            chart_data = _accounting_rollup_chart_rows(daily_rows, chart_axis_granularity)
        
        # Итоговые суммы
        total_income = sum(income_by_date.values())  # Уже чистый доход с налогом
        total_expected_income = sum(expected_income_by_date.values())
        total_expense = sum(expenses_by_date.values())
        net_profit = total_income - total_expense
        
        return {
            "total_income": total_income,
            "total_expected_income": total_expected_income,
            "total_expense": total_expense,
            "net_profit": net_profit,
            "total_points_spent": total_points_spent,  # Общая сумма списанных баллов в рублях
            "chart_data": chart_data,
            "chart_axis_granularity": chart_axis_granularity,
            "period": period,
            "start_date": start_date,
            "end_date": end_date
        }
    except Exception as e:
        logger.error(f"Ошибка при получении сводки: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/pending-confirmations")
async def get_pending_confirmations(
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Получить список услуг, ожидающих подтверждения (POST-visit, без finance-guard).
    Согласовано с POST /confirm-booking и с UI «Прошедшие записи» (bookingOutcome):
    прошедшие CREATED / CONFIRMED / AWAITING_CONFIRMATION без строки BookingConfirmation.
    Раньше здесь был только AWAITING_CONFIRMATION в БД — тогда прошлые CREATED/CONFIRMED
    показывались в списке с кнопкой, но не попадали в этот блок (ложный «рассинхрон»).
    При manual_confirm_enabled_at: только записи, созданные после включения ручного режима."""
    try:
        master_row_id = get_master_id_from_user(current_user.id, db)
        master = db.query(Master).filter(Master.id == master_row_id).first()
        master_id, indie_id = get_booking_owner_ids(current_user.id, db)
        owner_cond = _booking_owner_filter(Booking, master_id, indie_id)

        now_utc = datetime.utcnow()
        base_filters = [
            owner_cond,
            Booking.status.in_(
                [
                    BookingStatus.CREATED,
                    BookingStatus.CONFIRMED,
                    BookingStatus.AWAITING_CONFIRMATION,
                ]
            ),
            BookingConfirmation.id == None,  # Без подтверждения
            Booking.start_time < now_utc,  # POST-visit: только прошедшие
        ]
        # manual_confirm_enabled_at: показывать только записи, созданные после включения ручного режима
        manual_at = getattr(master, "manual_confirm_enabled_at", None) if master else None
        if manual_at is not None:
            base_filters.append(Booking.created_at >= manual_at)

        pending_bookings = db.query(Booking).outerjoin(
            BookingConfirmation,
            Booking.id == BookingConfirmation.booking_id
        ).filter(*base_filters).all()

        meta_map = {m.client_phone: m for m in db.query(MasterClientMetadata).filter(
            MasterClientMetadata.master_id == master_row_id
        ).all()}

        result = []
        for booking in pending_bookings:
            service = db.query(Service).filter(Service.id == booking.service_id).first()
            client = db.query(User).filter(User.id == booking.client_id).first()
            client_phone = client.phone if client and client.phone else None
            meta = get_meta_for_client(meta_map, client_phone)
            client_display_name = get_client_display_name(meta, client)
            svc_name = strip_indie_service_prefix(service.name if service else None) or "Услуга"
            result.append({
                "booking_id": booking.id,
                "client_name": client_display_name,
                "client_display_name": client_display_name,
                "client_phone": client_phone,
                "service_name": svc_name,
                "date": booking.start_time.date(),
                "start_time": booking.start_time,
                "payment_amount": booking.payment_amount,
                "status": booking.status,
            })
            logger.info(
                f"[pending-confirmations] master={current_user.id} booking_id={booking.id} "
                f"start_time={booking.start_time} status={booking.status} is_past={booking.start_time < now_utc}"
            )
        
        logger.info(f"Найдено {len(result)} услуг на подтверждение для мастера {current_user.id}")
        
        return {
            "pending_confirmations": result,
            "count": len(result)
        }
    except Exception as e:
        logger.error(f"Ошибка при получении услуг на подтверждение: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/confirm-booking/{booking_id}")
async def confirm_booking(
    booking_id: int,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Подтвердить завершение услуги и начислить доход/расходы (POST-visit, без finance-guard).
    Принимает CREATED и AWAITING_CONFIRMATION (как cancel-booking).
    """
    try:
        master_row_id = get_master_id_from_user(current_user.id, db)
        master_user_id = current_user.id
        master_id, indie_id = get_booking_owner_ids(current_user.id, db)
        owner_cond = _booking_owner_filter(Booking, master_id, indie_id)
        
        # Проверяем, что бронирование принадлежит мастеру и в допустимом статусе.
        # Post-visit confirm: CREATED, CONFIRMED (pre-visit confirmed), AWAITING_CONFIRMATION.
        booking = db.query(Booking).filter(
            Booking.id == booking_id,
            owner_cond,
            Booking.status.in_([BookingStatus.CREATED, BookingStatus.CONFIRMED, BookingStatus.AWAITING_CONFIRMATION])
        ).first()
        
        if not booking:
            # Может быть уже COMPLETED — идемпотентный успех
            existing = db.query(Booking).filter(
                Booking.id == booking_id,
                owner_cond,
            ).first()
            if existing and existing.status == BookingStatus.COMPLETED:
                conf = db.query(BookingConfirmation).filter(
                    BookingConfirmation.booking_id == booking_id
                ).first()
                if conf:
                    return {
                        "message": "Услуга уже подтверждена",
                        "confirmed_income": conf.confirmed_income,
                    }
            raise HTTPException(status_code=404, detail="Бронирование не найдено или не может быть подтверждено")
        
        # Будущие записи нельзя подтверждать
        if booking.start_time > datetime.utcnow():
            raise HTTPException(
                status_code=400,
                detail="Нельзя подтвердить будущую запись. Подтверждение доступно после времени начала.",
            )
        
        # Проверяем, не подтверждено ли уже
        existing_confirmation = db.query(BookingConfirmation).filter(
            BookingConfirmation.booking_id == booking_id
        ).first()
        
        if existing_confirmation:
            return {
                "message": "Услуга уже подтверждена",
                "confirmed_income": existing_confirmation.confirmed_income,
            }
        
        # Обработка баллов лояльности: списание и начисление
        points_spent = 0
        if booking.loyalty_points_used and booking.loyalty_points_used > 0:
            # Списываем баллы
            from utils.loyalty import spend_points
            try:
                spend_points(
                    db=db,
                    master_id=master_row_id,
                    client_id=booking.client_id,
                    amount=float(booking.loyalty_points_used),
                    booking_id=booking_id
                )
                points_spent = booking.loyalty_points_used
                logger.info(f"Списано {points_spent} баллов для записи {booking_id}")
            except Exception as e:
                logger.error(f"Ошибка при списании баллов для записи {booking_id}: {e}")
                # Не прерываем процесс подтверждения, но логируем ошибку
        
        # Вычисляем доход мастера (с учетом списанных баллов)
        actual_payment_amount = (booking.payment_amount or 0) - (points_spent or 0)
        
        # Создаем подтверждение
        confirmation = BookingConfirmation(
            booking_id=booking_id,
            master_id=master_user_id,
            confirmed_income=actual_payment_amount
        )
        db.add(confirmation)
        
        # Создаем income запись (доход уменьшается на сумму списанных баллов)
        income = Income(
            # Income.indie_master_id -> FK на indie_masters.id (не masters.id)
            indie_master_id=None,
            booking_id=booking_id,
            total_amount=booking.payment_amount or 0,  # Сумма к оплате после скидки
            master_earnings=actual_payment_amount,  # Фактический доход мастера
            salon_earnings=0,  # Для индивидуальных мастеров
            income_date=datetime.utcnow().date(),
            service_date=booking.start_time.date()
        )
        db.add(income)
        
        # Начисляем баллы (только с суммы, оплаченной деньгами)
        if booking.client_id:
            from utils.loyalty import get_loyalty_settings, earn_points
            loyalty_settings = get_loyalty_settings(db, master_row_id)
            
            if loyalty_settings and loyalty_settings.is_enabled and loyalty_settings.accrual_percent:
                # Баллы начисляются только с суммы, оплаченной деньгами
                amount_for_points = actual_payment_amount
                points_to_earn = int(amount_for_points * (loyalty_settings.accrual_percent / 100))
                
                if points_to_earn > 0:
                    try:
                        earn_points(
                            db=db,
                            master_id=master_row_id,
                            client_id=booking.client_id,
                            amount=points_to_earn,
                            booking_id=booking_id,
                            service_id=booking.service_id,
                            lifetime_days=loyalty_settings.points_lifetime_days
                        )
                        logger.info(f"Начислено {points_to_earn} баллов для записи {booking_id}")
                    except Exception as e:
                        logger.error(f"Ошибка при начислении баллов для записи {booking_id}: {e}")
        
        # Создаем расходы типа service_based для этой услуги
        service_expenses = db.query(MasterExpense).filter(
            MasterExpense.master_id == master_user_id,
            MasterExpense.expense_type == "service_based",
            MasterExpense.service_id == booking.service_id,
            MasterExpense.is_active == True
        ).all()
        
        for template in service_expenses:
            expense_record = MasterExpense(
                master_id=master_user_id,
                name=f"{template.name} (услуга #{booking_id})",
                expense_type="one_time",  # Создаем как разовый
                amount=template.amount,
                expense_date=datetime.utcnow()
            )
            db.add(expense_record)
        
        # Обновляем статус бронирования на completed
        booking.status = BookingStatus.COMPLETED
        
        db.commit()
        
        logger.info(f"Подтверждена услуга {booking_id} для мастера {current_user.id}, списано баллов: {points_spent}")
        
        return {
            "message": "Услуга успешно подтверждена",
            "confirmed_income": actual_payment_amount,
            "points_spent": points_spent,
            "expenses_created": len(service_expenses)
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка при подтверждении услуги: {e}")
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/confirm-all")
async def confirm_all_bookings(
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Подтвердить все неподтвержденные услуги (POST-visit, без finance-guard)."""
    try:
        master_row_id = get_master_id_from_user(current_user.id, db)
        master_user_id = current_user.id
        master_id, indie_id = get_booking_owner_ids(current_user.id, db)
        owner_cond = _booking_owner_filter(Booking, master_id, indie_id)
        
        # Получаем все confirmed бронирования в прошлом без подтверждения
        pending_bookings = db.query(Booking).outerjoin(
            BookingConfirmation,
            Booking.id == BookingConfirmation.booking_id
        ).filter(
            owner_cond,
            Booking.status == BookingStatus.AWAITING_CONFIRMATION,
            BookingConfirmation.id == None
        ).all()
        
        confirmed_count = 0
        for booking in pending_bookings:
            # Создаем подтверждение
            confirmation = BookingConfirmation(
                booking_id=booking.id,
                master_id=master_user_id,
                confirmed_income=booking.payment_amount or 0
            )
            db.add(confirmation)
            
            # Обновляем статус бронирования на completed
            booking.status = BookingStatus.COMPLETED
            
            confirmed_count += 1
        
        db.commit()
        
        logger.info(f"Подтверждено {confirmed_count} услуг для мастера {current_user.id}")
        
        return {
            "message": f"Подтверждено {confirmed_count} услуг",
            "confirmed_count": confirmed_count
        }
    except Exception as e:
        logger.error(f"Ошибка при массовом подтверждении услуг: {e}")
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/cancel-booking/{booking_id}")
async def cancel_booking(
    booking_id: int,
    cancellation_reason: str = Query(..., regex="^(client_requested|client_no_show|mutual_agreement|master_unavailable)$"),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Отклонить отдельную услугу (POST-visit, без finance-guard)."""
    try:
        master_id, indie_id = get_booking_owner_ids(current_user.id, db)
        owner_cond = _booking_owner_filter(Booking, master_id, indie_id)
        
        # Проверяем, что бронирование принадлежит мастеру
        booking = db.query(Booking).filter(
            Booking.id == booking_id,
            owner_cond,
            Booking.status.in_([BookingStatus.CREATED, BookingStatus.CONFIRMED, BookingStatus.AWAITING_CONFIRMATION])
        ).first()
        
        if not booking:
            raise HTTPException(status_code=404, detail="Бронирование не найдено или не может быть отклонено")
        
        # Проверяем, не подтверждено ли уже
        existing_confirmation = db.query(BookingConfirmation).filter(
            BookingConfirmation.booking_id == booking_id
        ).first()
        
        if existing_confirmation:
            raise HTTPException(status_code=400, detail="Услуга уже подтверждена")
        
        # При отмене сбрасываем резервирование баллов (баллы не были списаны, так как запись не подтверждена)
        if booking.loyalty_points_used:
            booking.loyalty_points_used = 0
        
        # Устанавливаем причину отмены и инициатора
        booking.cancelled_by_user_id = current_user.id
        booking.cancellation_reason = cancellation_reason
        
        # Обновляем статус бронирования на cancelled
        booking.status = BookingStatus.CANCELLED
        
        db.commit()
        
        logger.info(f"Отклонена услуга {booking_id} для мастера {current_user.id}")
        
        return {
            "message": "Услуга успешно отклонена",
            "booking_id": booking_id
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка при отклонении услуги {booking_id}: {e}")
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/cancel-all")
async def cancel_all_bookings(
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Отменить все неподтвержденные услуги (POST-visit, без finance-guard)."""
    try:
        master_id, indie_id = get_booking_owner_ids(current_user.id, db)
        owner_cond = _booking_owner_filter(Booking, master_id, indie_id)
        
        # Получаем все confirmed бронирования в прошлом без подтверждения
        pending_bookings = db.query(Booking).outerjoin(
            BookingConfirmation,
            Booking.id == BookingConfirmation.booking_id
        ).filter(
            owner_cond,
            Booking.status == BookingStatus.AWAITING_CONFIRMATION,
            BookingConfirmation.id == None
        ).all()
        
        cancelled_count = 0
        for booking in pending_bookings:
            # Обновляем статус бронирования на cancelled
            booking.status = BookingStatus.CANCELLED
            cancelled_count += 1
        
        db.commit()
        
        logger.info(f"Отменено {cancelled_count} услуг для мастера {current_user.id}")
        
        return {
            "message": f"Отменено {cancelled_count} услуг",
            "cancelled_count": cancelled_count
        }
    except Exception as e:
        logger.error(f"Ошибка при массовой отмене услуг: {e}")
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/export")
async def export_data(
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    format: str = Query("csv", regex="^(csv|excel)$"),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Экспорт данных бухгалтерии в CSV или Excel"""
    _ensure_finance_access(db, current_user.id)
    try:
        # Получаем данные
        if not start_date:
            start_date = datetime.now() - timedelta(days=30)
        if not end_date:
            end_date = datetime.now()
        
        # Доходы
        incomes = db.query(BookingConfirmation).filter(
            BookingConfirmation.master_id == current_user.id,
            BookingConfirmation.confirmed_at >= start_date,
            BookingConfirmation.confirmed_at <= end_date
        ).all()
        
        # Расходы
        expenses = db.query(MasterExpense).filter(
            MasterExpense.master_id == current_user.id,
            MasterExpense.expense_date >= start_date,
            MasterExpense.expense_date <= end_date
        ).all()
        
        # Формируем данные для экспорта
        rows = []
        
        for income in incomes:
            rows.append({
                "Дата": income.confirmed_at.strftime("%Y-%m-%d %H:%M"),
                "Тип операции": "Доход",
                "Название": f"Услуга #{income.booking_id}",
                "Доход": income.confirmed_income,
                "Расход": 0,
            })
        
        for expense in expenses:
            rows.append({
                "Дата": expense.expense_date.strftime("%Y-%m-%d %H:%M") if expense.expense_date else "",
                "Тип операции": "Расход",
                "Название": expense.name,
                "Доход": 0,
                "Расход": expense.amount,
            })
        
        # Сортируем по дате
        rows.sort(key=lambda x: x["Дата"], reverse=True)
        
        # Добавляем баланс
        balance = 0
        for row in reversed(rows):
            balance += row["Доход"] - row["Расход"]
            row["Баланс"] = balance
        
        if format == "csv":
            # Создаем CSV
            output = BytesIO()
            output.write('\ufeff'.encode('utf-8'))  # BOM для корректного отображения кириллицы в Excel
            
            fieldnames = ["Дата", "Тип операции", "Название", "Доход", "Расход", "Баланс"]
            writer = csv.DictWriter(output, fieldnames=fieldnames, delimiter=';')
            writer.writeheader()
            writer.writerows(rows)
            
            output.seek(0)
            
            return Response(
                content=output.getvalue(),
                media_type="text/csv; charset=utf-8",
                headers={
                    "Content-Disposition": f"attachment; filename=accounting_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.csv"
                }
            )
        else:
            # Создаем Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Бухгалтерия"
            
            # Заголовки
            headers = ["Дата", "Тип операции", "Название", "Доход", "Расход", "Баланс"]
            ws.append(headers)
            
            # Стилизация заголовков
            header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Данные
            for row in rows:
                ws.append([
                    row["Дата"],
                    row["Тип операции"],
                    row["Название"],
                    row["Доход"],
                    row["Расход"],
                    row["Баланс"]
                ])
            
            # Автоматическая ширина колонок
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Форматирование чисел
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                # Доход (колонка 4)
                if row[3].value and row[3].value != 0:
                    row[3].font = Font(color="4CAF50")
                # Расход (колонка 5)
                if row[4].value and row[4].value != 0:
                    row[4].font = Font(color="F44336")
                # Баланс (колонка 6)
                if row[5].value:
                    if row[5].value >= 0:
                        row[5].font = Font(color="2196F3", bold=True)
                    else:
                        row[5].font = Font(color="F44336", bold=True)
            
            # Сохраняем в BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            return Response(
                content=output.getvalue(),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f"attachment; filename=accounting_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
                }
            )
            
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка при экспорте данных: {e}")
        raise HTTPException(status_code=500, detail=str(e))

